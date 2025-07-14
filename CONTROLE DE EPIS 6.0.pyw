import customtkinter as ctk
from tkinter import ttk, Listbox, END, filedialog
from PIL import Image, ImageTk
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from tkinter import messagebox
import os
from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkcalendar import DateEntry
import locale
import webbrowser
from io import BytesIO
import barcode
from barcode.writer import ImageWriter
import unicodedata

# --- CORREÇÃO DEFINITIVA PARA O ERRO DE LOCALE ---
try:
    locale.setlocale(locale.LC_NUMERIC, 'C')
    locale.setlocale(locale.LC_MONETARY, 'pt_BR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_MONETARY, 'Portuguese_Brazil.1252')
    except locale.Error:
        print("Aviso: Não foi possível definir a localidade monetária brasileira.")
# --- FIM DA CORREÇÃO ---


# --- Configurações da Interface ---
APP_TITLE = "Controller EPI"
WINDOW_SIZE = "1200x750"
BACKGROUND_COLOR = "#FFFFFF"
SIDEBAR_COLOR = "#F0F0F0"
RECENT_ENTRIES_BG_COLOR = "#F8F9FA"
CARD_BORDER_COLOR = "#E0E0E0"
BUTTON_COLOR = "#2E8B57" # Verde
BUTTON_HOVER_COLOR = "#3CB371"
BUTTON_COLOR_RED = "#D2042D" # Vermelho (Scarlet)
BUTTON_HOVER_RED = "#FF3333" # Vermelho Claro
BUTTON_COLOR_YELLOW = "#FFC107" # Amarelo
BUTTON_HOVER_YELLOW = "#FFCA28"
BUTTON_COLOR_BLUE = "#2196F3" # Azul
BUTTON_HOVER_BLUE = "#42A5F5"
INFO_BG_COLOR = "#E8F5E9" # Verde claro
INFO_TEXT_COLOR = "#1B5E20" # Verde escuro
INFO_ICON_COLOR = "#2E8B57" # Verde padrão
TEXT_COLOR = "#000000"
CORNER_RADIUS = 6

# --- Fontes ---
FONT_FAMILY = "Roboto"
TITLE_FONT = (FONT_FAMILY, 26, "bold")
SUBTITLE_FONT = (FONT_FAMILY, 16, "bold")
INPUT_FONT = (FONT_FAMILY, 13)
LABEL_FONT = (FONT_FAMILY, 14)
BUTTON_FONT = (FONT_FAMILY, 14, "bold")
FOOTER_FONT = (FONT_FAMILY, 13)
INFO_FONT = (FONT_FAMILY, 14, "italic")
VERSION_FONT = (FONT_FAMILY, 11)

# --- Constantes de Layout ---
ENTRY_WIDTH = 320

# --- Caminhos ---
BASE_DIR = r"G:\Meu Drive\CONTROLLER\DATA CENTER"
DB_PATH = os.path.join(BASE_DIR, "BANCO_DE_DADOS_EPI.xlsx")
LOGO_PATH = os.path.join(BASE_DIR, "LOGO_RUBBERGATTI.png")
WORKSHEET_NAME = "CONTROLE EPI"
CADASTRO_WORKSHEET_NAME = "CADASTRO EPI"


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE); self.configure(fg_color=BACKGROUND_COLOR); self.minsize(1100, 700)
        
        width = self.winfo_screenwidth()
        height = self.winfo_screenheight()
        self.geometry(f"{width}x{height}+0+0")
        
        self.chart_canvas = None
        self.last_added_row_index = None
        self.last_movimentacao_row_index = None
        self.selected_excel_row = None
        self.sort_by_date_asc = True
        
        self.employee_list = []
        self.epi_list = []
        self.epi_ca_map = {}
        self.epi_price_map = {}
        self.last_hovered_suggestion = -1
        self.updates_window = None
        self.drilldown_employee = None

        self.grid_columnconfigure(1, weight=1); self.grid_rowconfigure(0, weight=1)
        self.create_sidebar()
        
        self.content_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.content_frame.grid(row=0, column=1, sticky="nsew", padx=30, pady=20)
        self.content_frame.grid_columnconfigure(0, weight=1); self.content_frame.grid_rowconfigure(0, weight=1)
        
        self.controle_epi_frame = self.create_controle_epi_screen()
        self.cadastro_epi_frame = self.create_cadastro_epi_screen()
        self.cadastro_geral_epi_frame = self.create_cadastro_geral_epi_screen()
        self.inventario_epi_frame = self.create_inventario_epi_screen()
        self.dashboard_frame = self.create_dashboard_screen()
        
        self.bind("<Escape>", self.handle_escape_key)
        
        self.select_frame_by_name("controle_epi")

    def create_sidebar(self):
        self.sidebar_frame = ctk.CTkFrame(self, width=200, corner_radius=0, fg_color=SIDEBAR_COLOR)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(7, weight=1)

        logo_frame = ctk.CTkFrame(self.sidebar_frame, fg_color="white", corner_radius=CORNER_RADIUS)
        logo_frame.grid(row=0, column=0, padx=20, pady=(20, 5))
        
        try:
            pil_image = Image.open(LOGO_PATH)
            original_width, original_height = pil_image.size
            target_width = 180
            aspect_ratio = original_height / original_width
            new_height = int(target_width * aspect_ratio)
            
            logo_image = ctk.CTkImage(light_image=pil_image, size=(target_width, new_height))
            ctk.CTkLabel(logo_frame, text="", image=logo_image).pack(padx=5, pady=5)
        except FileNotFoundError:
            ctk.CTkLabel(logo_frame, text="Logo não encontrada", font=INFO_FONT).pack(padx=10, pady=10)

        ctk.CTkLabel(self.sidebar_frame, text="Rubber Guardian", font=ctk.CTkFont(family=TITLE_FONT[0], size=18, weight="bold")).grid(row=1, column=0, pady=(0, 20))

        self.controle_button = ctk.CTkButton(self.sidebar_frame, text="Controle EPI", command=self.controle_button_event, font=BUTTON_FONT, corner_radius=CORNER_RADIUS); self.controle_button.grid(row=2, column=0, padx=20, pady=4, sticky="ew")
        self.cadastro_button = ctk.CTkButton(self.sidebar_frame, text="Entrada / Ajuste", command=self.cadastro_button_event, font=BUTTON_FONT, corner_radius=CORNER_RADIUS); self.cadastro_button.grid(row=3, column=0, padx=20, pady=4, sticky="ew")
        
        self.cadastro_geral_button = ctk.CTkButton(self.sidebar_frame, text="Cadastro EPI", command=self.cadastro_geral_button_event, font=BUTTON_FONT, corner_radius=CORNER_RADIUS)
        self.cadastro_geral_button.grid(row=4, column=0, padx=20, pady=4, sticky="ew")

        self.inventario_button = ctk.CTkButton(self.sidebar_frame, text="Inventário de EPIs", command=self.inventario_button_event, font=BUTTON_FONT, corner_radius=CORNER_RADIUS); self.inventario_button.grid(row=5, column=0, padx=20, pady=4, sticky="ew")
        self.dashboard_button = ctk.CTkButton(self.sidebar_frame, text="Dashboard", command=self.dashboard_button_event, font=BUTTON_FONT, corner_radius=CORNER_RADIUS); self.dashboard_button.grid(row=6, column=0, padx=20, pady=4, sticky="ew")
        
        self.updates_button = ctk.CTkButton(self.sidebar_frame, text="Atualizações", command=self.open_updates_window, font=BUTTON_FONT, corner_radius=CORNER_RADIUS); self.updates_button.grid(row=8, column=0, padx=20, pady=10, sticky="ew")
        ctk.CTkLabel(self.sidebar_frame, text="V5.0.7", font=VERSION_FONT).grid(row=9, column=0, padx=20, pady=(10, 20), sticky="s")

    def create_form_row(self, parent, label_text, row_index, widget_class=ctk.CTkEntry, **kwargs):
        label = ctk.CTkLabel(parent, text=label_text, font=LABEL_FONT)
        label.grid(row=row_index, column=0, sticky="w", padx=(0, 25))
        
        entry = widget_class(parent, width=ENTRY_WIDTH, height=35, font=INPUT_FONT, corner_radius=CORNER_RADIUS, **kwargs)
        entry.grid(row=row_index, column=1, sticky="w", pady=5)
        return entry

    def create_controle_epi_screen(self):
        main_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(2, weight=1)
        main_frame.grid_rowconfigure(1, weight=1)

        form_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        form_frame.grid(row=0, column=0, rowspan=2, sticky="nw")
        form_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(form_frame, text="Lançamento de Entrega de EPI", font=TITLE_FONT).grid(row=0, column=0, columnspan=2, pady=(0, 20), sticky="w")
        
        info_frame = ctk.CTkFrame(form_frame, fg_color=INFO_BG_COLOR, corner_radius=CORNER_RADIUS)
        info_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 25))
        ctk.CTkLabel(info_frame, text="ⓘ", font=ctk.CTkFont(size=20), text_color=INFO_ICON_COLOR).pack(side="left", padx=(10, 5))
        ctk.CTkLabel(info_frame, text="Preencha os dados abaixo para registrar a entrega do EPI ao funcionário.", font=INFO_FONT, text_color=INFO_TEXT_COLOR).pack(side="left", padx=(0, 10), pady=10)

        self.funcionario_entry = self.create_form_row(form_frame, "Nome do Funcionário:", 2, placeholder_text="Ex: João da Silva")
        self.suggestion_listbox = Listbox(form_frame, font=INPUT_FONT, bg="#EAEAEA", bd=0, highlightthickness=0, selectbackground=BUTTON_COLOR, selectforeground="white")
        self.funcionario_entry.bind("<KeyRelease>", self.update_suggestions)
        self.suggestion_listbox.bind("<<ListboxSelect>>", self.select_suggestion)
        self.suggestion_listbox.bind("<Motion>", self.on_suggestion_hover)
        self.suggestion_listbox.bind("<Leave>", self.on_suggestion_leave)
        self.funcionario_entry.bind("<FocusOut>", lambda e: self.suggestion_listbox.place_forget())

        self.data_saida_entry = self.create_form_row(form_frame, "Data da Entrega:", 3)
        self.data_saida_entry.insert(0, datetime.now().strftime("%d/%m/%Y"))

        self.quantidade_saida_entry = self.create_form_row(form_frame, "Quantidade Entregue:", 4, placeholder_text="Ex: 1")

        self.nome_epi_saida_entry = self.create_form_row(form_frame, "Nome do Equipamento/EPI:", 5, widget_class=ctk.CTkComboBox, values=[], command=self.on_epi_select)
        self.nome_epi_saida_entry.set('')
        self.nome_epi_saida_entry.bind("<KeyRelease>", self.on_epi_type)

        self.ca_saida_entry = self.create_form_row(form_frame, "C.A:", 6, placeholder_text="Será preenchido automaticamente")

        self.nome_epi_saida_entry.bind("<Return>", lambda event: self.quantidade_saida_entry.focus_set())
        self.quantidade_saida_entry.bind("<Return>", lambda event: self.add_saida_data())

        action_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        action_frame.grid(row=7, column=0, columnspan=2, pady=(40, 10), sticky="w")
        add_button = ctk.CTkButton(action_frame, text="➕ Adicionar Registro", command=self.add_saida_data, fg_color=BUTTON_COLOR, hover_color=BUTTON_HOVER_COLOR, font=BUTTON_FONT, height=40, width=ENTRY_WIDTH, corner_radius=CORNER_RADIUS)
        add_button.pack(side="left")
        self.remove_button = ctk.CTkButton(action_frame, text="Remover Lançamento Recente", command=self.remove_specific_entry, fg_color=BUTTON_COLOR_RED, hover_color=BUTTON_HOVER_RED, font=BUTTON_FONT, height=40, corner_radius=CORNER_RADIUS)

        separator = ttk.Separator(main_frame, orient='vertical')
        separator.grid(row=0, column=1, rowspan=2, sticky='ns', padx=20)
        
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Recent.Treeview", background=RECENT_ENTRIES_BG_COLOR, fieldbackground=RECENT_ENTRIES_BG_COLOR, foreground=TEXT_COLOR, rowheight=25, font=INPUT_FONT)
        style.configure("Recent.Treeview.Heading", font=ctk.CTkFont(family="Roboto", size=12, weight="bold"))
        style.layout("Recent.Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])

        saldo_geral_frame = ctk.CTkFrame(main_frame, fg_color=RECENT_ENTRIES_BG_COLOR, corner_radius=CORNER_RADIUS, border_width=1, border_color=CARD_BORDER_COLOR)
        saldo_geral_frame.grid(row=0, column=2, sticky="nsew", pady=(0, 20))
        saldo_geral_frame.grid_columnconfigure(0, weight=1)
        saldo_geral_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(saldo_geral_frame, text="Saldo Geral de EPIs", font=SUBTITLE_FONT).grid(row=0, column=0, pady=(20,10))

        self.saldo_geral_tree = ttk.Treeview(saldo_geral_frame, style="Recent.Treeview", columns=("EPI", "Saldo", "Valor"), show="headings")
        self.saldo_geral_tree.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0,15))

        self.saldo_geral_tree.heading("EPI", text="Nome do EPI", anchor="center")
        self.saldo_geral_tree.heading("Saldo", text="Saldo em Estoque", anchor="center")
        self.saldo_geral_tree.heading("Valor", text="Valor em Estoque", anchor="center")

        self.saldo_geral_tree.column("EPI", stretch=True, anchor="w")
        self.saldo_geral_tree.column("Saldo", width=120, anchor="center")
        self.saldo_geral_tree.column("Valor", width=120, anchor="center")

        recent_frame = ctk.CTkFrame(main_frame, fg_color=RECENT_ENTRIES_BG_COLOR, corner_radius=CORNER_RADIUS, border_width=1, border_color=CARD_BORDER_COLOR)
        recent_frame.grid(row=1, column=2, sticky="nsew")
        recent_frame.grid_columnconfigure(0, weight=1)
        recent_frame.grid_rowconfigure(1, weight=1)
        
        title_label = ctk.CTkLabel(recent_frame, text="Ultimos Registros de Entrega", font=SUBTITLE_FONT)
        title_label.grid(row=0, column=0, pady=(20,10))
        
        self.recent_entries_tree = ttk.Treeview(recent_frame, style="Recent.Treeview", columns=("Func", "EPI", "Data", "Qtd"), show="headings")
        self.recent_entries_tree.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0,15))
        
        self.recent_entries_tree.heading("Func", text="Funcionário")
        self.recent_entries_tree.heading("EPI", text="EPI")
        self.recent_entries_tree.heading("Data", text="Data")
        self.recent_entries_tree.heading("Qtd", text="Qtd")
        
        self.recent_entries_tree.column("Func", width=150, stretch=True)
        self.recent_entries_tree.column("EPI", width=150, stretch=True)
        self.recent_entries_tree.column("Data", width=80, anchor="center")
        self.recent_entries_tree.column("Qtd", width=40, anchor="center")

        return main_frame

    def create_cadastro_epi_screen(self):
        main_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(2, weight=1)

        form_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        form_frame.grid(row=0, column=0, sticky="nw")
        form_frame.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(form_frame, text="Entrada e Ajuste de Estoque", font=TITLE_FONT).grid(row=0, column=0, columnspan=2, pady=(0, 40), sticky="w")
        
        info_frame = ctk.CTkFrame(form_frame, fg_color=INFO_BG_COLOR, corner_radius=CORNER_RADIUS)
        info_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 25))
        ctk.CTkLabel(info_frame, text="ⓘ", font=ctk.CTkFont(size=20), text_color=INFO_ICON_COLOR).pack(side="left", padx=(10, 5))
        ctk.CTkLabel(info_frame, text="Preencha os dados para registrar uma nova entrada ou ajuste no estoque.", font=INFO_FONT, text_color=INFO_TEXT_COLOR).pack(side="left", padx=(0, 10), pady=10)

        self.mov_type_entry = self.create_form_row(form_frame, "Tipo de Movimentação:", 2, widget_class=ctk.CTkComboBox, values=["ENTRADA", "AJUSTE"])
        self.mov_nome_epi_entry = self.create_form_row(form_frame, "Nome EPI:", 3, widget_class=ctk.CTkComboBox, values=[], command=self.on_mov_epi_select)
        self.mov_nome_epi_entry.set('')
        # --- INÍCIO DA IMPLEMENTAÇÃO: Binding para busca inteligente ---
        self.mov_nome_epi_entry.bind("<KeyRelease>", self.on_mov_epi_type)
        # --- FIM DA IMPLEMENTAÇÃO ---
        self.mov_ca_entry = self.create_form_row(form_frame, "C.A:", 4)
        self.mov_data_entrada_entry = self.create_form_row(form_frame, "Data Entrada:", 5)
        self.mov_data_entrada_entry.insert(0, datetime.now().strftime("%d/%m/%Y"))
        self.mov_qtd_entrada_entry = self.create_form_row(form_frame, "Qtd. Entrada:", 6)
        
        self.mov_nome_epi_entry.bind("<Return>", lambda event: self.mov_qtd_entrada_entry.focus_set())
        self.mov_qtd_entrada_entry.bind("<Return>", lambda event: self.add_movimentacao_data())

        action_button_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        action_button_frame.grid(row=7, column=0, columnspan=2, pady=(20,0), sticky="w")
        
        self.add_mov_button = ctk.CTkButton(action_button_frame, text="➕ Adicionar Movimentação", command=self.add_movimentacao_data, fg_color=BUTTON_COLOR, hover_color=BUTTON_HOVER_COLOR, font=BUTTON_FONT, height=40, width=ENTRY_WIDTH, corner_radius=CORNER_RADIUS)
        self.add_mov_button.pack(side="left", padx=(0, 10))
        
        self.remove_mov_button = ctk.CTkButton(action_button_frame, text="Remover Recente", command=self.remove_recent_movimentacao, font=BUTTON_FONT, height=40, fg_color=BUTTON_COLOR_RED, hover_color=BUTTON_HOVER_RED, corner_radius=CORNER_RADIUS)
        
        separator = ttk.Separator(main_frame, orient='vertical')
        separator.grid(row=0, column=1, sticky='ns', padx=20)

        recent_mov_frame = ctk.CTkFrame(main_frame, fg_color=RECENT_ENTRIES_BG_COLOR, corner_radius=CORNER_RADIUS, border_width=1, border_color=CARD_BORDER_COLOR)
        recent_mov_frame.grid(row=0, column=2, sticky="nsew")
        recent_mov_frame.grid_columnconfigure(0, weight=1)
        recent_mov_frame.grid_rowconfigure(1, weight=1)
        
        ctk.CTkLabel(recent_mov_frame, text="Últimas Movimentações", font=SUBTITLE_FONT).grid(row=0, column=0, pady=(20,10))
        
        self.recent_mov_tree = ttk.Treeview(recent_mov_frame, style="Recent.Treeview", columns=("Mov", "EPI", "Data", "Qtd"), show="headings")
        self.recent_mov_tree.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0,15))
        
        self.recent_mov_tree.heading("Mov", text="Movimentação", anchor="center")
        self.recent_mov_tree.heading("EPI", text="EPI", anchor="center")
        self.recent_mov_tree.heading("Data", text="Data", anchor="center")
        self.recent_mov_tree.heading("Qtd", text="Qtd", anchor="center")
        
        self.recent_mov_tree.column("Mov", width=100, stretch=True, anchor="center")
        self.recent_mov_tree.column("EPI", width=150, stretch=True, anchor="center")
        self.recent_mov_tree.column("Data", width=80, anchor="center")
        self.recent_mov_tree.column("Qtd", width=40, anchor="center")

        return main_frame

    def create_cadastro_geral_epi_screen(self):
        main_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(2, weight=1)

        form_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        form_frame.grid(row=0, column=0, sticky="nw")
        form_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(form_frame, text="Cadastro de Tipos de EPI", font=TITLE_FONT).grid(row=0, column=0, columnspan=2, pady=(0, 20), sticky="w")
        
        self.cadastro_geral_nome_entry = self.create_form_row(form_frame, "Nome do EPI:", 1, placeholder_text="Ex: Luva de Algodão")
        self.cadastro_geral_ca_entry = self.create_form_row(form_frame, "C.A:", 2, placeholder_text="Ex: 37930")
        self.cadastro_geral_preco_entry = self.create_form_row(form_frame, "Preço Unitário (R$):", 3, placeholder_text="Ex: 2,00")

        self.cadastro_geral_nome_entry.bind("<Return>", lambda event: self.cadastro_geral_ca_entry.focus_set())
        self.cadastro_geral_ca_entry.bind("<Return>", lambda event: self.cadastro_geral_preco_entry.focus_set())
        self.cadastro_geral_preco_entry.bind("<Return>", lambda event: self.add_cadastro_epi_data())

        action_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        action_frame.grid(row=4, column=0, columnspan=2, pady=(40, 10), sticky="w")
        add_button = ctk.CTkButton(action_frame, text="➕ Cadastrar EPI", command=self.add_cadastro_epi_data, fg_color=BUTTON_COLOR, hover_color=BUTTON_HOVER_COLOR, font=BUTTON_FONT, height=40, width=ENTRY_WIDTH, corner_radius=CORNER_RADIUS)
        add_button.pack(side="left")

        separator = ttk.Separator(main_frame, orient='vertical')
        separator.grid(row=0, column=1, sticky='ns', padx=20)

        list_frame = ctk.CTkFrame(main_frame, fg_color=RECENT_ENTRIES_BG_COLOR, corner_radius=CORNER_RADIUS, border_width=1, border_color=CARD_BORDER_COLOR)
        list_frame.grid(row=0, column=2, sticky="nsew")
        list_frame.grid_columnconfigure(0, weight=1)
        list_frame.grid_rowconfigure(1, weight=1)
        
        ctk.CTkLabel(list_frame, text="EPIs Cadastrados", font=SUBTITLE_FONT).grid(row=0, column=0, pady=(20,10))
        
        self.cadastro_geral_tree = ttk.Treeview(list_frame, style="Recent.Treeview", columns=("EPI", "CA", "Preço"), show="headings")
        self.cadastro_geral_tree.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0,15))
        
        self.cadastro_geral_tree.heading("EPI", text="Nome do EPI", anchor="center")
        self.cadastro_geral_tree.heading("CA", text="C.A", anchor="center")
        self.cadastro_geral_tree.heading("Preço", text="Preço Unitário", anchor="center")
        
        self.cadastro_geral_tree.column("EPI", width=200, stretch=True, anchor="center")
        self.cadastro_geral_tree.column("CA", width=100, anchor="center")
        self.cadastro_geral_tree.column("Preço", width=100, anchor="center")

        self.cadastro_geral_tree.bind('<Double-1>', self.open_epi_edit_popup)

        return main_frame

    def create_inventario_epi_screen(self):
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        frame.grid_columnconfigure(0, weight=1); frame.grid_rowconfigure(2, weight=1)
        
        top_frame = ctk.CTkFrame(frame, fg_color="transparent"); top_frame.grid(row=0, column=0, sticky="ew", pady=(0, 15))
        ctk.CTkLabel(top_frame, text="Inventário de EPIs", font=TITLE_FONT).pack(side="left")
        
        filter_controls_frame = ctk.CTkFrame(frame, fg_color="transparent")
        filter_controls_frame.grid(row=1, column=0, sticky="ew", pady=(0, 15))
        
        ctk.CTkLabel(filter_controls_frame, text="EPI:", font=LABEL_FONT).pack(side="left", padx=(0,5))
        self.epi_filter = ctk.CTkComboBox(filter_controls_frame, values=["Todos"], width=200, font=INPUT_FONT, corner_radius=CORNER_RADIUS); self.epi_filter.pack(side="left", padx=(0,15))
        
        ctk.CTkLabel(filter_controls_frame, text="C.A:", font=LABEL_FONT).pack(side="left", padx=(0,5))
        self.ca_filter = ctk.CTkComboBox(filter_controls_frame, values=["Todos"], width=100, font=INPUT_FONT, corner_radius=CORNER_RADIUS); self.ca_filter.pack(side="left", padx=(0,15))
        
        ctk.CTkLabel(filter_controls_frame, text="Funcionário:", font=LABEL_FONT).pack(side="left", padx=(0,5))
        self.func_filter = ctk.CTkComboBox(filter_controls_frame, values=["Todos"], width=200, font=INPUT_FONT, corner_radius=CORNER_RADIUS); self.func_filter.pack(side="left", padx=(0,15))
        
        ctk.CTkButton(filter_controls_frame, text="Aplicar", command=self.populate_treeview, height=35, font=BUTTON_FONT, corner_radius=CORNER_RADIUS).pack(side="left", padx=(10,5))
        ctk.CTkButton(filter_controls_frame, text="Limpar", command=self.reset_inventory_filters, height=35, font=BUTTON_FONT, corner_radius=CORNER_RADIUS).pack(side="left")

        tree_frame = ctk.CTkFrame(frame, fg_color="transparent"); tree_frame.grid(row=2, column=0, sticky="nsew"); tree_frame.grid_columnconfigure(0, weight=1); tree_frame.grid_rowconfigure(0, weight=1)
        
        style = ttk.Style(); style.theme_use("default")
        style.configure("Treeview", background="#FFFFFF", fieldbackground="#FFFFFF", foreground="#000000", rowheight=25, font=INPUT_FONT)
        style.map('Treeview', background=[('selected', BUTTON_COLOR)]); style.configure("Treeview.Heading", font=ctk.CTkFont(family="Roboto", size=13, weight="bold"))
        
        self.tree = ttk.Treeview(tree_frame, columns=("EPI", "CA", "Func", "Data Saída", "Qtd Saída", "Data Entrada", "Qtd Entrada", "Preço Unitário"), show='headings')
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview); vsb.grid(row=0, column=1, sticky="ns"); self.tree.configure(yscrollcommand=vsb.set)
        
        headings = {"EPI": 200, "CA": 80, "Func": 150, "Data Saída": 100, "Qtd Saída": 80, "Data Entrada": 100, "Qtd Entrada": 80, "Preço Unitário": 100}
        for col, width in headings.items(): self.tree.heading(col, text=col); self.tree.column(col, width=width, anchor="center")
        self.tree.bind('<<TreeviewSelect>>', self.on_item_select)
        
        self.tree.bind('<Double-1>', self.open_edit_popup)

        bottom_frame = ctk.CTkFrame(frame, fg_color="transparent")
        bottom_frame.grid(row=3, column=0, sticky="ew", pady=(15,0))
        bottom_frame.grid_columnconfigure(2, weight=1)
        
        self.saldo_frame = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        self.saldo_frame.grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(self.saldo_frame, text="Saldo Atual:", font=LABEL_FONT).pack(side="left")
        self.saldo_label = ctk.CTkLabel(self.saldo_frame, text="-", font=ctk.CTkFont(family=INPUT_FONT[0], size=18, weight="bold"))
        self.saldo_label.pack(side="left", padx=10)
        
        self.valor_estoque_frame = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        self.valor_estoque_frame.grid(row=0, column=1, sticky="w", padx=20)
        ctk.CTkLabel(self.valor_estoque_frame, text="Valor de Estoque:", font=LABEL_FONT).pack(side="left")
        self.valor_estoque_label = ctk.CTkLabel(self.valor_estoque_frame, text="-", font=ctk.CTkFont(family=INPUT_FONT[0], size=18, weight="bold"))
        self.valor_estoque_label.pack(side="left", padx=10)
        
        self.remove_from_inventory_button = ctk.CTkButton(bottom_frame, text="Remover Registro Selecionado", command=self.delete_record, fg_color=BUTTON_COLOR_RED, hover_color=BUTTON_HOVER_RED, font=BUTTON_FONT, height=40, corner_radius=CORNER_RADIUS)
        self.remove_from_inventory_button.grid(row=0, column=2, sticky="e")
        
        return frame

    def create_dashboard_screen(self):
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(2, weight=1)

        ctk.CTkLabel(frame, text="Dashboard de Consumo de EPIs", font=TITLE_FONT).grid(row=0, column=0, pady=(0, 10), sticky="w")

        controls_main_frame = ctk.CTkFrame(frame, fg_color="transparent")
        controls_main_frame.grid(row=1, column=0, sticky="ew", pady=(0, 20))
        controls_main_frame.grid_columnconfigure(1, weight=1)

        top_row_frame = ctk.CTkFrame(controls_main_frame, fg_color="transparent")
        top_row_frame.grid(row=0, column=0, columnspan=3, sticky="ew")
        top_row_frame.grid_columnconfigure(1, weight=1)

        left_filters_frame = ctk.CTkFrame(top_row_frame, fg_color="transparent")
        left_filters_frame.grid(row=0, column=0, sticky="w")

        ctk.CTkLabel(left_filters_frame, text="Tipo de Gráfico:", font=LABEL_FONT).pack(side="left", padx=(0, 10))
        self.chart_filter = ctk.CTkComboBox(left_filters_frame, values=["Quantidade Total de EPIs", "Frequência", "EPIs"], command=self.update_dashboard_controls, font=INPUT_FONT, width=200, corner_radius=CORNER_RADIUS)
        self.chart_filter.pack(side="left", padx=(0, 20))

        ctk.CTkLabel(left_filters_frame, text="Visualizar como:", font=LABEL_FONT).pack(side="left", padx=(0, 10))
        self.view_mode_filter = ctk.CTkComboBox(left_filters_frame, values=["📊 Quantidade", "💸 Valor", "📊+💸 Qtd e Valor"], command=self.apply_filters, font=INPUT_FONT, width=200, corner_radius=CORNER_RADIUS)
        self.view_mode_filter.pack(side="left", padx=(0, 20))

        ctk.CTkLabel(left_filters_frame, text="Filtrar por EPI:", font=LABEL_FONT).pack(side="left", padx=(0, 10))
        self.dashboard_epi_filter = ctk.CTkComboBox(left_filters_frame, values=["Todos"], font=INPUT_FONT, width=200, command=self.apply_filters, corner_radius=CORNER_RADIUS)
        self.dashboard_epi_filter.pack(side="left")

        right_buttons_frame = ctk.CTkFrame(top_row_frame, fg_color="transparent")
        right_buttons_frame.grid(row=0, column=2, sticky="e")

        ctk.CTkButton(right_buttons_frame, text="Aplicar", command=self.apply_filters, height=35, font=BUTTON_FONT, corner_radius=CORNER_RADIUS, fg_color=BUTTON_COLOR, hover_color=BUTTON_HOVER_COLOR).pack(side="left", padx=(10, 5))
        ctk.CTkButton(right_buttons_frame, text="Limpar", command=self.clear_all_dashboard_filters, height=35, font=BUTTON_FONT, corner_radius=CORNER_RADIUS, fg_color=BUTTON_COLOR_YELLOW, hover_color=BUTTON_HOVER_YELLOW).pack(side="left", padx=5)
        ctk.CTkButton(right_buttons_frame, text="Imprimir", command=self.print_placeholder, height=35, font=BUTTON_FONT, corner_radius=CORNER_RADIUS, fg_color=BUTTON_COLOR_BLUE, hover_color=BUTTON_HOVER_BLUE).pack(side="left", padx=5)

        date_frame = ctk.CTkFrame(controls_main_frame, fg_color="transparent")
        date_frame.grid(row=1, column=0, columnspan=3, sticky="w", pady=(10, 0))

        ctk.CTkLabel(date_frame, text="Data Início:", font=LABEL_FONT).pack(side="left", padx=(0, 10))
        self.start_date_cal = DateEntry(date_frame, width=12, date_pattern='dd/mm/yyyy', font=INPUT_FONT)
        self.start_date_cal.pack(side="left", padx=(0, 20))
        
        ctk.CTkLabel(date_frame, text="Data Fim:", font=LABEL_FONT).pack(side="left", padx=(0, 10))
        self.end_date_cal = DateEntry(date_frame, width=12, date_pattern='dd/mm/yyyy', font=INPUT_FONT)
        self.end_date_cal.pack(side="left")

        self.chart_frame = ctk.CTkFrame(frame, fg_color="transparent")
        self.chart_frame.grid(row=2, column=0, sticky="nsew")

        totals_frame = ctk.CTkFrame(frame, fg_color="transparent")
        totals_frame.grid(row=3, column=0, pady=15)
        self.sum_label_title = ctk.CTkLabel(totals_frame, text="Somatório Período:", font=LABEL_FONT)
        self.sum_label_title.grid(row=0, column=0, padx=(0,10))
        self.sum_label = ctk.CTkLabel(totals_frame, text="-", font=FOOTER_FONT)
        self.sum_label.grid(row=0, column=1)
        self.mean_label_title = ctk.CTkLabel(totals_frame, text="Média Período:", font=LABEL_FONT)
        self.mean_label_title.grid(row=0, column=2, padx=(50,10))
        self.mean_label = ctk.CTkLabel(totals_frame, text="-", font=FOOTER_FONT)
        self.mean_label.grid(row=0, column=3)
        
        return frame

    def print_placeholder(self):
        messagebox.showinfo("Informação", "A funcionalidade de impressão será implementada em uma versão futura.")

    def open_updates_window(self):
        if self.updates_window is not None and self.updates_window.winfo_exists():
            self.updates_window.focus()
            return

        self.updates_window = ctk.CTkToplevel(self)
        self.updates_window.title("Histórico de Alterações")
        self.updates_window.geometry("700x550")
        self.updates_window.resizable(False, False)
        self.updates_window.transient(self)
        self.updates_window.grab_set()

        self.updates_window.grid_columnconfigure(0, weight=1)
        self.updates_window.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(self.updates_window, text="Controller EPI - Histórico de Versões", font=ctk.CTkFont(family=TITLE_FONT[0], size=22)).grid(row=0, column=0, pady=20)

        textbox = ctk.CTkTextbox(self.updates_window, font=INPUT_FONT, wrap="word", corner_radius=CORNER_RADIUS)
        textbox.grid(row=1, column=0, sticky="nsew", padx=20)
        
        changelog_text = """
## Versão 5.0.7 (Atual)
- Implementada busca inteligente e accent-insensitive para os campos de EPI, corrigindo o preenchimento do C.A.

## Versão 5.0.6
- Adicionada função para retornar ao gráfico principal do dashboard com a tecla ESC.
- Corrigido erro na geração de código de barras para nomes com acentos (ex: "LUVA DE ALGODÃO").
- Corrigida a formatação do número C.A na tabela de EPIs Cadastrados para incluir separador de milhar.

## Versão 5.0.5
- Adicionado painel "Saldo Geral de EPIs" na tela de Lançamento de Entrega, com cálculo de saldo e valor em estoque.
- O novo painel é atualizado dinamicamente após cada movimentação.
- Invertida a ordem dos painéis na tela de Lançamento para priorizar o Saldo Geral.

## Versão 5.0.4
- Corrigido o drill-down do dashboard: agora, ao clicar em um funcionário na visualização "Qtd e Valor", o gráfico de consumo mensal também exibe "Qtd e Valor" corretamente.
"""
        textbox.insert("1.0", changelog_text)
        textbox.configure(state="disabled")

        button_frame = ctk.CTkFrame(self.updates_window, fg_color="transparent")
        button_frame.grid(row=2, column=0, pady=20)
        
        ctk.CTkButton(button_frame, text="Suporte / Sobre", command=lambda: webbrowser.open("https://github.com/sanctuslocalhost"), font=BUTTON_FONT, corner_radius=CORNER_RADIUS).pack(side="left", padx=10)
        ctk.CTkButton(button_frame, text="Fechar", command=self.updates_window.destroy, font=BUTTON_FONT, corner_radius=CORNER_RADIUS).pack(side="left", padx=10)

    def handle_escape_key(self, event=None):
        if self.dashboard_frame.winfo_ismapped() and self.drilldown_employee:
            self.drilldown_employee = None
            self.apply_filters()

    def load_epi_prices(self):
        try:
            self.get_cadastro_epi_workbook()
            
            df_prices = pd.read_excel(DB_PATH, sheet_name=CADASTRO_WORKSHEET_NAME)
            
            df_prices.dropna(subset=['NOME EPI'], inplace=True)
            df_prices['NOME EPI'] = df_prices['NOME EPI'].astype(str).str.strip()
            
            if 'PREÇO UNITÁRIO' in df_prices.columns:
                df_prices['PREÇO UNITÁRIO'] = df_prices['PREÇO UNITÁRIO'].astype(str).str.replace('R$', '', regex=False).str.replace('.', '', regex=False).str.replace(',', '.', regex=False).str.strip()
                df_prices['PREÇO UNITÁRIO'] = pd.to_numeric(df_prices['PREÇO UNITÁRIO'], errors='coerce').fillna(0.0)
            else:
                df_prices['PREÇO UNITÁRIO'] = 0.0

            self.epi_price_map = pd.Series(df_prices['PREÇO UNITÁRIO'].values, index=df_prices['NOME EPI']).to_dict()

        except FileNotFoundError:
            self.epi_price_map = {}
        except ValueError:
            self.epi_price_map = {}
        except Exception as e:
            messagebox.showwarning("Aviso", f"Não foi possível carregar os preços dos EPIs:\n{e}")
            self.epi_price_map = {}

    def populate_controle_epi_options(self):
        try:
            self.load_epi_prices()
            
            self.epi_list = sorted(self.epi_price_map.keys())

            df = pd.read_excel(DB_PATH, sheet_name=WORKSHEET_NAME)
            df.dropna(subset=['FUNCIONARIO'], inplace=True)
            self.employee_list = sorted(df['FUNCIONARIO'].astype(str).str.strip().unique().tolist())
            
            df_epis = pd.read_excel(DB_PATH, sheet_name=CADASTRO_WORKSHEET_NAME)
            df_epis = df_epis.dropna(subset=['NOME EPI', 'C.A'])
            self.epi_ca_map = pd.Series(df_epis['C.A'].astype(str).str.strip().values, index=df_epis['NOME EPI'].astype(str).str.strip()).to_dict()
            
            self.nome_epi_saida_entry.configure(values=self.epi_list)
            self.mov_nome_epi_entry.configure(values=self.epi_list)
        except Exception as e:
            messagebox.showwarning("Aviso", f"Não foi possível carregar as listas de autocompletar:\n{e}")

    def update_recent_entries_panel(self):
        for i in self.recent_entries_tree.get_children():
            self.recent_entries_tree.delete(i)
        try:
            df = pd.read_excel(DB_PATH, sheet_name=WORKSHEET_NAME)
            df_saidas = df.dropna(subset=['DATA SAIDA', 'FUNCIONARIO', 'NOME EPI', 'QUANTIDADE SAIDA']).copy()
            
            df_saidas['DATA SAIDA'] = pd.to_datetime(df_saidas['DATA SAIDA'], dayfirst=True, errors='coerce')
            df_saidas.dropna(subset=['DATA SAIDA'], inplace=True)
            
            recent_entries = df_saidas.sort_values(by='DATA SAIDA', ascending=False).head(13)
            recent_entries_sorted = recent_entries.sort_values(by='DATA SAIDA', ascending=True)

            for _, row in recent_entries_sorted.iterrows():
                func = str(row['FUNCIONARIO'])
                epi = str(row['NOME EPI'])
                data = row['DATA SAIDA'].strftime('%d/%m/%y')
                qtd = int(row['QUANTIDADE SAIDA'])
                self.recent_entries_tree.insert("", "end", values=(func, epi, data, qtd))

        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível carregar registros recentes:\n{e}")

    def update_recent_movimentacoes_panel(self):
        for i in self.recent_mov_tree.get_children():
            self.recent_mov_tree.delete(i)
        try:
            df = pd.read_excel(DB_PATH, sheet_name=WORKSHEET_NAME)
            df_mov = df.dropna(subset=['DATA ENTRADA', 'QUANTIDADE ENTRADA', 'FUNCIONARIO', 'NOME EPI']).copy()
            
            df_mov['DATA ENTRADA'] = pd.to_datetime(df_mov['DATA ENTRADA'], dayfirst=True, errors='coerce')
            df_mov.dropna(subset=['DATA ENTRADA'], inplace=True)
            
            recent_movs = df_mov.sort_values(by='DATA ENTRADA', ascending=False).head(13)
            recent_movs_sorted = recent_movs.sort_values(by='DATA ENTRADA', ascending=True)

            for _, row in recent_movs_sorted.iterrows():
                mov = str(row['FUNCIONARIO'])
                epi = str(row['NOME EPI'])
                data = row['DATA ENTRADA'].strftime('%d/%m/%y')
                qtd = int(row['QUANTIDADE ENTRADA'])
                self.recent_mov_tree.insert("", "end", values=(mov, epi, data, qtd))

        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível carregar movimentações recentes:\n{e}")

    def populate_saldo_geral_treeview(self):
        for i in self.saldo_geral_tree.get_children():
            self.saldo_geral_tree.delete(i)
        
        try:
            df_controle = pd.read_excel(DB_PATH, sheet_name=WORKSHEET_NAME)
            df_cadastro = pd.read_excel(DB_PATH, sheet_name=CADASTRO_WORKSHEET_NAME)

            df_controle['QUANTIDADE ENTRADA'] = pd.to_numeric(df_controle['QUANTIDADE ENTRADA'], errors='coerce').fillna(0)
            df_controle['QUANTIDADE SAIDA'] = pd.to_numeric(df_controle['QUANTIDADE SAIDA'], errors='coerce').fillna(0)
            df_controle.dropna(subset=['NOME EPI'], inplace=True)
            df_controle['NOME EPI'] = df_controle['NOME EPI'].astype(str).str.strip()

            saldos = df_controle.groupby('NOME EPI').agg(
                total_entrada=('QUANTIDADE ENTRADA', 'sum'),
                total_saida=('QUANTIDADE SAIDA', 'sum')
            ).reset_index()
            saldos['Saldo'] = saldos['total_entrada'] - saldos['total_saida']

            df_cadastro.dropna(subset=['NOME EPI'], inplace=True)
            df_cadastro['NOME EPI'] = df_cadastro['NOME EPI'].astype(str).str.strip()
            if 'PREÇO UNITÁRIO' in df_cadastro.columns:
                df_cadastro['PREÇO UNITÁRIO'] = pd.to_numeric(
                    df_cadastro['PREÇO UNITÁRIO'].astype(str).str.replace('R$', '', regex=False).str.replace('.', '', regex=False).str.replace(',', '.', regex=False).str.strip(),
                    errors='coerce'
                ).fillna(0.0)
            else:
                df_cadastro['PREÇO UNITÁRIO'] = 0.0

            df_final = pd.merge(df_cadastro, saldos[['NOME EPI', 'Saldo']], on='NOME EPI', how='left')
            df_final['Saldo'] = df_final['Saldo'].fillna(0).astype(int)
            
            df_final['Valor em Estoque'] = df_final['Saldo'] * df_final['PREÇO UNITÁRIO']

            for _, row in df_final.iterrows():
                nome_epi = row['NOME EPI']
                saldo = row['Saldo']
                valor_estoque = locale.currency(row['Valor em Estoque'], grouping=True)
                self.saldo_geral_tree.insert("", "end", values=(nome_epi, saldo, valor_estoque))

        except Exception as e:
            if "No sheet named" not in str(e):
                messagebox.showerror("Erro ao Calcular Saldo", f"Não foi possível calcular o saldo geral:\n{e}")

    def update_suggestions(self, event):
        typed_text = self.funcionario_entry.get().lower()
        self.suggestion_listbox.delete(0, END)
        
        if not typed_text:
            self.suggestion_listbox.place_forget()
            return

        suggestions = [name for name in self.employee_list if typed_text in name.lower()]
        if suggestions:
            for name in suggestions:
                self.suggestion_listbox.insert(END, name)
            
            x = self.funcionario_entry.winfo_x()
            y = self.funcionario_entry.winfo_y() + self.funcionario_entry.winfo_height()
            width = self.funcionario_entry.winfo_width()
            
            list_height = min(len(suggestions) * 28, 140)
            self.suggestion_listbox.place(x=x, y=y, width=width, height=list_height)
            self.suggestion_listbox.lift()
        else:
            self.suggestion_listbox.place_forget()

    def select_suggestion(self, event):
        if self.suggestion_listbox.curselection():
            selected_name = self.suggestion_listbox.get(self.suggestion_listbox.curselection())
            self.funcionario_entry.delete(0, END)
            self.funcionario_entry.insert(0, selected_name)
            self.suggestion_listbox.place_forget()
            self.funcionario_entry.focus()

    def on_suggestion_hover(self, event):
        listbox = event.widget
        new_hovered_index = listbox.nearest(event.y)
        
        if self.last_hovered_suggestion != new_hovered_index:
            if self.last_hovered_suggestion != -1:
                listbox.itemconfig(self.last_hovered_suggestion, {'bg': '#EAEAEA', 'fg': TEXT_COLOR})
            
            listbox.itemconfig(new_hovered_index, {'bg': BUTTON_HOVER_COLOR, 'fg': 'white'})
            self.last_hovered_suggestion = new_hovered_index

    def on_suggestion_leave(self, event):
        if self.last_hovered_suggestion != -1:
            event.widget.itemconfig(self.last_hovered_suggestion, {'bg': '#EAEAEA', 'fg': TEXT_COLOR})
            self.last_hovered_suggestion = -1

    # --- INÍCIO DA IMPLEMENTAÇÃO: Funções de busca inteligente ---
    def _normalize_string(self, text):
        """Remove acentos e converte para minúsculas para comparação."""
        if not isinstance(text, str):
            return ""
        return unicodedata.normalize('NFD', text).encode('ascii', 'ignore').decode('utf-8').lower()

    def _find_ca_for_epi(self, epi_name_to_find):
        """Encontra o C.A. para um nome de EPI, ignorando acentos e maiúsculas/minúsculas."""
        normalized_to_find = self._normalize_string(epi_name_to_find)
        if not normalized_to_find:
            return None

        for original_epi_name, ca_value in self.epi_ca_map.items():
            normalized_original = self._normalize_string(original_epi_name)
            if normalized_original == normalized_to_find:
                return ca_value
        return None

    def on_epi_select(self, selected_epi):
        self.ca_saida_entry.delete(0, END)
        ca_value = self._find_ca_for_epi(selected_epi)
        if ca_value:
            self.ca_saida_entry.insert(0, self.clean_cell_value(ca_value, is_ca=True))

    def on_epi_type(self, event):
        typed_epi = self.nome_epi_saida_entry.get()
        self.on_epi_select(typed_epi)
        
    def on_mov_epi_select(self, selected_epi):
        self.mov_ca_entry.delete(0, END)
        ca_value = self._find_ca_for_epi(selected_epi)
        if ca_value:
            self.mov_ca_entry.insert(0, self.clean_cell_value(ca_value, is_ca=True))

    def on_mov_epi_type(self, event):
        typed_epi = self.mov_nome_epi_entry.get()
        self.on_mov_epi_select(typed_epi)
    # --- FIM DA IMPLEMENTAÇÃO ---

    def get_workbook(self):
        if not os.path.exists(BASE_DIR): os.makedirs(BASE_DIR)
        if not os.path.exists(DB_PATH):
            workbook = openpyxl.Workbook(); sheet = workbook.active; sheet.title = WORKSHEET_NAME
            headers = ["NOME EPI", "C.A", "FUNCIONARIO", "DATA SAIDA", "QUANTIDADE SAIDA", "DATA ENTRADA", "QUANTIDADE ENTRADA", "PREÇO UNITÁRIO"]
            sheet.append(headers); workbook.save(DB_PATH)
        return load_workbook(DB_PATH)

    def get_cadastro_epi_workbook(self):
        workbook = self.get_workbook()
        if CADASTRO_WORKSHEET_NAME not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=CADASTRO_WORKSHEET_NAME)
            headers = ["NOME EPI", "C.A", "PREÇO UNITÁRIO"]
            sheet.append(headers)
            workbook.save(DB_PATH)
        return workbook

    def format_new_row(self, sheet, row_index):
        center_alignment = Alignment(horizontal='center', vertical='center'); thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in range(1, 9): cell = sheet.cell(row=row_index, column=col); cell.alignment = center_alignment; cell.border = thin_border

    def clean_cell_value(self, value, is_numeric=False, is_ca=False, is_currency=False):
        if pd.isna(value): return ""
        if isinstance(value, datetime): return value.strftime("%d/%m/%Y")
        if is_numeric:
            try: return f"{int(float(value))}"
            except (ValueError, TypeError): return ""
        if is_ca:
            try: return f"{int(float(value)):,}".replace(",", ".")
            except (ValueError, TypeError): return str(value)
        if is_currency:
            try: return locale.currency(float(value), grouping=True)
            except (ValueError, TypeError): return ""
        return str(value)

    def parse_date_for_sort(self, date_str):
        if not isinstance(date_str, str) or not date_str: return datetime.max
        try: return datetime.strptime(date_str, "%d/%m/%Y")
        except ValueError: return datetime.max

    def add_saida_data(self):
        campos = [self.funcionario_entry.get().strip(), self.data_saida_entry.get().strip(), self.quantidade_saida_entry.get().strip(), self.nome_epi_saida_entry.get().strip(), self.ca_saida_entry.get().strip()]
        if not all(campos): messagebox.showwarning("Atenção", "Todos os campos são obrigatórios."); return
        try:
            nome_epi = campos[3]
            preco_unitario = self.epi_price_map.get(nome_epi, 0.0)

            nova_linha = [nome_epi, campos[4], campos[0], self.clean_cell_value(datetime.strptime(campos[1], "%d/%m/%Y")), int(campos[2]), "", "", preco_unitario]
            workbook = self.get_workbook(); sheet = workbook[WORKSHEET_NAME]
            sheet.append(nova_linha); self.last_added_row_index = sheet.max_row; self.format_new_row(sheet, self.last_added_row_index); workbook.save(DB_PATH)
            messagebox.showinfo("Sucesso", f"Entrega para '{campos[0]}' registrada!"); self.remove_button.pack(side="left", padx=(10, 0))
            
            self.funcionario_entry.delete(0, 'end'); self.quantidade_saida_entry.delete(0, 'end'); self.nome_epi_saida_entry.set(''); self.ca_saida_entry.delete(0, 'end')
            self.populate_controle_epi_options()
            self.update_recent_entries_panel()
            self.populate_saldo_geral_treeview()

        except ValueError: messagebox.showwarning("Atenção", "Verifique o formato da data (DD/MM/AAAA) e se a quantidade é um número.")
        except PermissionError: messagebox.showerror("Erro de Permissão", f"Não foi possível salvar o arquivo.\n\nPor favor, feche a planilha '{os.path.basename(DB_PATH)}' e tente novamente.")
        except Exception as e: messagebox.showerror("Erro", f"Ocorreu um erro inesperado ao salvar:\n{e}")

    def add_movimentacao_data(self):
        campos_obrigatorios = [self.mov_type_entry.get(), self.mov_nome_epi_entry.get(), self.mov_data_entrada_entry.get(), self.mov_qtd_entrada_entry.get()]
        if not all(campos_obrigatorios): messagebox.showwarning("Atenção", "Todos os campos, exceto C.A., são obrigatórios."); return
        
        try:
            mov_type = self.mov_type_entry.get()
            nome_epi = self.mov_nome_epi_entry.get()
            ca = self.mov_ca_entry.get()
            data_entrada = self.mov_data_entrada_entry.get()
            qtd_entrada = self.mov_qtd_entrada_entry.get()

            preco_unitario = self.epi_price_map.get(nome_epi, 0.0)

            nova_linha = [nome_epi, ca, mov_type, "", "", self.clean_cell_value(datetime.strptime(data_entrada, "%d/%m/%Y")), int(qtd_entrada), preco_unitario]
            workbook = self.get_workbook(); sheet = workbook[WORKSHEET_NAME]
            sheet.append(nova_linha); self.last_movimentacao_row_index = sheet.max_row; self.format_new_row(sheet, self.last_movimentacao_row_index); workbook.save(DB_PATH)
            messagebox.showinfo("Sucesso", f"Movimentação '{mov_type}' para o EPI '{nome_epi}' registrada!"); self.remove_mov_button.pack(side="left", padx=(10, 0))
            
            self.mov_nome_epi_entry.set(''); self.mov_ca_entry.delete(0, 'end'); self.mov_qtd_entrada_entry.delete(0, 'end')
            self.populate_controle_epi_options()
            self.update_recent_movimentacoes_panel()
            self.populate_saldo_geral_treeview()

        except ValueError: messagebox.showwarning("Atenção", "Verifique o formato da data (DD/MM/AAAA) e se a quantidade é um número.")
        except PermissionError: messagebox.showerror("Erro de Permissão", f"Não foi possível salvar o arquivo.\n\nPor favor, feche a planilha '{os.path.basename(DB_PATH)}' e tente novamente.")
        except Exception as e: messagebox.showerror("Erro", f"Ocorreu um erro inesperado ao salvar:\n{e}")

    def add_cadastro_epi_data(self):
        nome = self.cadastro_geral_nome_entry.get().strip()
        ca = self.cadastro_geral_ca_entry.get().strip()
        preco_str = self.cadastro_geral_preco_entry.get().strip().replace(",", ".")

        if not all([nome, ca, preco_str]):
            messagebox.showwarning("Atenção", "Todos os campos são obrigatórios.")
            return
        
        try:
            preco = float(preco_str)
            
            workbook = self.get_cadastro_epi_workbook()
            sheet = workbook[CADASTRO_WORKSHEET_NAME]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[0].strip().lower() == nome.lower():
                    messagebox.showwarning("Atenção", f"O EPI '{nome}' já está cadastrado.")
                    return

            sheet.append([nome, ca, preco])
            workbook.save(DB_PATH)
            
            messagebox.showinfo("Sucesso", f"EPI '{nome}' cadastrado com sucesso!")
            self.cadastro_geral_nome_entry.delete(0, 'end')
            self.cadastro_geral_ca_entry.delete(0, 'end')
            self.cadastro_geral_preco_entry.delete(0, 'end')
            
            self.populate_cadastro_geral_treeview()
            self.load_epi_prices()

        except ValueError:
            messagebox.showwarning("Atenção", "O preço deve ser um número válido (ex: 2.50 ou 2,50).")
        except PermissionError:
            messagebox.showerror("Erro de Permissão", f"Não foi possível salvar o arquivo.\n\nPor favor, feche a planilha '{os.path.basename(DB_PATH)}' e tente novamente.")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro inesperado ao salvar:\n{e}")

    def remove_specific_entry(self):
        if self.last_added_row_index is None: messagebox.showwarning("Aviso", "Nenhum registro recente foi adicionado para remover."); return
        if not messagebox.askyesno("Confirmar Remoção", "Deseja remover o registro que acabou de adicionar?"): return
        try:
            workbook = self.get_workbook(); sheet = workbook[WORKSHEET_NAME]; log_info = f"EPI: {sheet.cell(row=self.last_added_row_index, column=1).value or ''}"
            sheet.delete_rows(self.last_added_row_index); workbook.save(DB_PATH)
            messagebox.showinfo("Sucesso", f"O registro recente foi removido.\n({log_info})"); self.remove_button.pack_forget(); self.last_added_row_index = None
            self.update_recent_entries_panel()
            self.populate_saldo_geral_treeview()
        except PermissionError: messagebox.showerror("Erro de Permissão", f"Não foi possível remover o registro.\n\nPor favor, feche a planilha '{os.path.basename(DB_PATH)}' e tente novamente.")
        except Exception as e: messagebox.showerror("Erro", f"Ocorreu um erro ao remover:\n{e}")

    def remove_recent_movimentacao(self):
        if self.last_movimentacao_row_index is None: messagebox.showwarning("Aviso", "Nenhuma movimentação recente foi adicionada para remover."); return
        if not messagebox.askyesno("Confirmar Remoção", "Deseja remover a movimentação que acabou de adicionar?"): return
        try:
            workbook = self.get_workbook(); sheet = workbook[WORKSHEET_NAME]; log_info = f"EPI: {sheet.cell(row=self.last_movimentacao_row_index, column=1).value or ''}"
            sheet.delete_rows(self.last_movimentacao_row_index); workbook.save(DB_PATH)
            messagebox.showinfo("Sucesso", f"A movimentação recente foi removida.\n({log_info})"); self.remove_mov_button.pack_forget(); self.last_movimentacao_row_index = None
            self.update_recent_movimentacoes_panel()
            self.populate_saldo_geral_treeview()
        except PermissionError: messagebox.showerror("Erro de Permissão", f"Não foi possível remover o registro.\n\nPor favor, feche a planilha '{os.path.basename(DB_PATH)}' e tente novamente.")
        except Exception as e: messagebox.showerror("Erro", f"Ocorreu um erro ao remover:\n{e}")

    def reset_screens(self):
        self.selected_excel_row = None
        if hasattr(self, 'remove_button'): self.remove_button.pack_forget()
        if hasattr(self, 'remove_mov_button'): self.remove_mov_button.pack_forget()
        if hasattr(self, 'remove_from_inventory_button'): self.remove_from_inventory_button.configure(state="disabled")
        if hasattr(self, 'chart_canvas') and self.chart_canvas: self.chart_canvas.get_tk_widget().destroy(); self.chart_canvas = None
        if hasattr(self, 'sum_label'): self.sum_label.configure(text="-")
        if hasattr(self, 'mean_label'): self.mean_label.configure(text="-")
        if hasattr(self, 'saldo_frame'): self.saldo_frame.grid_forget()
        if hasattr(self, 'valor_estoque_frame'): self.valor_estoque_frame.grid_forget()

    def populate_treeview(self):
        try:
            self.tree.delete(*self.tree.get_children())
            df = pd.read_excel(DB_PATH, sheet_name=WORKSHEET_NAME)
            
            if 'epi_filter' in self.__dict__:
                self.epi_filter.configure(values=["Todos"] + sorted(df['NOME EPI'].dropna().unique().tolist()))
                self.ca_filter.configure(values=["Todos"] + sorted(df['C.A'].dropna().astype(str).unique().tolist()))
                self.func_filter.configure(values=["Todos"] + sorted(df['FUNCIONARIO'].dropna().unique().tolist()))

            epi, ca, func = self.epi_filter.get(), self.ca_filter.get(), self.func_filter.get()
            filtered_df = df.copy()
            if epi != "Todos": filtered_df = filtered_df[filtered_df['NOME EPI'] == epi]
            if ca != "Todos": filtered_df = filtered_df[filtered_df['C.A'].astype(str) == ca]
            if func != "Todos": filtered_df = filtered_df[filtered_df['FUNCIONARIO'] == func]

            filtered_df['sort_date'] = filtered_df.apply(lambda row: self.parse_date_for_sort(self.clean_cell_value(row.iloc[3]) or self.clean_cell_value(row.iloc[5])), axis=1)
            filtered_df.sort_values(by='sort_date', ascending=self.sort_by_date_asc, inplace=True)

            for index, row in filtered_df.iterrows():
                display_values = (
                    self.clean_cell_value(row.iloc[0]), self.clean_cell_value(row.iloc[1], is_ca=True),
                    self.clean_cell_value(row.iloc[2]), self.clean_cell_value(row.iloc[3]),
                    self.clean_cell_value(row.iloc[4], is_numeric=True), self.clean_cell_value(row.iloc[5]),
                    self.clean_cell_value(row.iloc[6], is_numeric=True), self.clean_cell_value(row.iloc[7], is_currency=True)
                )
                self.tree.insert('', 'end', iid=row.name + 2, values=display_values)
            
            if epi != "Todos" or ca != "Todos":
                total_entrada = pd.to_numeric(filtered_df['QUANTIDADE ENTRADA'], errors='coerce').fillna(0).sum()
                total_saida = pd.to_numeric(filtered_df['QUANTIDADE SAIDA'], errors='coerce').fillna(0).sum()
                saldo = total_entrada - total_saida
                self.saldo_label.configure(text=f"{int(saldo)}")
                self.saldo_frame.grid(row=0, column=0, sticky="w")
                
                filtered_df['PREÇO UNITÁRIO'] = pd.to_numeric(filtered_df['PREÇO UNITÁRIO'], errors='coerce').fillna(0)
                valor_total_estoque = (filtered_df['QUANTIDADE ENTRADA'] * filtered_df['PREÇO UNITÁRIO']).sum() - (filtered_df['QUANTIDADE SAIDA'] * filtered_df['PREÇO UNITÁRIO']).sum()
                self.valor_estoque_label.configure(text=locale.currency(valor_total_estoque, grouping=True))
                self.valor_estoque_frame.grid(row=0, column=1, sticky="w", padx=20)
            else:
                self.saldo_frame.grid_forget()
                self.valor_estoque_frame.grid_forget()

        except Exception as e: messagebox.showerror("Erro", f"Não foi possível carregar a lista de EPIs:\n{e}")

    def populate_cadastro_geral_treeview(self):
        for i in self.cadastro_geral_tree.get_children():
            self.cadastro_geral_tree.delete(i)
        try:
            self.get_cadastro_epi_workbook()
            df = pd.read_excel(DB_PATH, sheet_name=CADASTRO_WORKSHEET_NAME)
            df.dropna(how='all', inplace=True)

            for index, row in df.iterrows():
                nome = self.clean_cell_value(row.iloc[0])
                ca = self.clean_cell_value(row.iloc[1], is_ca=True)
                preco = self.clean_cell_value(row.iloc[2], is_currency=True)
                self.cadastro_geral_tree.insert("", "end", iid=index + 2, values=(nome, ca, preco))
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível carregar a lista de EPIs cadastrados:\n{e}")

    def reset_inventory_filters(self):
        self.epi_filter.set("Todos"); self.ca_filter.set("Todos"); self.func_filter.set("Todos")
        self.populate_treeview()

    def toggle_date_sort(self):
        self.sort_by_date_asc = not self.sort_by_date_asc
        self.populate_treeview()

    def on_item_select(self, event):
        if not self.tree.selection(): return
        self.selected_excel_row = int(self.tree.selection()[0])
        self.remove_from_inventory_button.configure(state="normal")

    def delete_record(self):
        if self.selected_excel_row is None: messagebox.showerror("Erro", "Nenhum registro selecionado para remover."); return
        if not messagebox.askyesno("Confirmar Remoção", "TEM CERTEZA que deseja remover este registro permanentemente?"): return
        try:
            workbook = self.get_workbook(); sheet = workbook[WORKSHEET_NAME]
            sheet.delete_rows(self.selected_excel_row); workbook.save(DB_PATH); messagebox.showinfo("Sucesso", "O registro foi removido.")
            self.populate_treeview()
        except PermissionError: messagebox.showerror("Erro de Permissão", f"Não foi possível remover o registro.\n\nPor favor, feche a planilha '{os.path.basename(DB_PATH)}' e tente novamente.")
        except Exception as e: messagebox.showerror("Erro ao Remover", f"Não foi possível remover:\n{e}")

    def open_edit_popup(self, event):
        selection = self.tree.selection()
        if not selection:
            return
        
        selected_item = selection[0]
        excel_row_index = int(selected_item)
        values = self.tree.item(selected_item, 'values')

        is_saida_record = bool(values[3]) and bool(values[4])

        edit_window = ctk.CTkToplevel(self)
        edit_window.transient(self)
        edit_window.grab_set()
        edit_window.resizable(False, False)
        edit_window.configure(fg_color=BACKGROUND_COLOR)

        form_frame = ctk.CTkFrame(edit_window, fg_color="transparent")
        form_frame.pack(padx=30, pady=20)

        if is_saida_record:
            edit_window.title("Editar Lançamento de Entrega")
            title_label = ctk.CTkLabel(form_frame, text="Editar Lançamento de Entrega", font=SUBTITLE_FONT)
            title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20), sticky="ew")
            
            nome_epi_entry = self.create_form_row(form_frame, "Nome do Equipamento/EPI:", 1)
            ca_entry = self.create_form_row(form_frame, "C.A:", 2)
            func_entry = self.create_form_row(form_frame, "Nome do Funcionário:", 3)
            data_saida_entry = self.create_form_row(form_frame, "Data da Entrega:", 4)
            qtd_saida_entry = self.create_form_row(form_frame, "Quantidade Entregue:", 5)

            nome_epi_entry.insert(0, values[0])
            ca_entry.insert(0, values[1])
            func_entry.insert(0, values[2])
            data_saida_entry.insert(0, values[3])
            qtd_saida_entry.insert(0, values[4])

            save_command = lambda: self.save_edited_record(
                excel_row_index,
                {
                    'nome_epi': nome_epi_entry.get(), 'ca': ca_entry.get(), 'funcionario': func_entry.get(),
                    'data_saida': data_saida_entry.get(), 'qtd_saida': qtd_saida_entry.get()
                },
                'saida', edit_window
            )

        else:
            edit_window.title("Editar Movimentação de Estoque")
            title_label = ctk.CTkLabel(form_frame, text="Editar Movimentação de Estoque", font=SUBTITLE_FONT)
            title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20), sticky="ew")

            mov_type_entry = self.create_form_row(form_frame, "Tipo de Movimentação:", 1, widget_class=ctk.CTkComboBox, values=["ENTRADA", "AJUSTE"])
            nome_epi_entry = self.create_form_row(form_frame, "Nome EPI:", 2)
            ca_entry = self.create_form_row(form_frame, "C.A:", 3)
            data_entrada_entry = self.create_form_row(form_frame, "Data Entrada:", 4)
            qtd_entrada_entry = self.create_form_row(form_frame, "Qtd. Entrada:", 5)

            mov_type_entry.set(values[2])
            nome_epi_entry.insert(0, values[0])
            ca_entry.insert(0, values[1])
            data_entrada_entry.insert(0, values[5])
            qtd_entrada_entry.insert(0, values[6])

            save_command = lambda: self.save_edited_record(
                excel_row_index,
                {
                    'mov_type': mov_type_entry.get(), 'nome_epi': nome_epi_entry.get(), 'ca': ca_entry.get(),
                    'data_entrada': data_entrada_entry.get(), 'qtd_entrada': qtd_entrada_entry.get()
                },
                'entrada', edit_window
            )

        action_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        action_frame.grid(row=6, column=0, columnspan=2, pady=(30, 10), sticky="ew")
        action_frame.grid_columnconfigure(0, weight=1)
        action_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkButton(action_frame, text="Salvar Alterações", command=save_command, font=BUTTON_FONT, fg_color=BUTTON_COLOR, hover_color=BUTTON_HOVER_COLOR).grid(row=0, column=0, padx=5, sticky="e")
        ctk.CTkButton(action_frame, text="Cancelar", command=edit_window.destroy, font=BUTTON_FONT, fg_color=BUTTON_COLOR_RED, hover_color=BUTTON_HOVER_RED).grid(row=0, column=1, padx=5, sticky="w")

    def save_edited_record(self, excel_row_index, new_data, record_type, popup_window):
        try:
            workbook = self.get_workbook()
            sheet = workbook[WORKSHEET_NAME]

            if record_type == 'saida':
                if not all([new_data['nome_epi'], new_data['funcionario'], new_data['data_saida'], new_data['qtd_saida']]):
                    messagebox.showwarning("Atenção", "Todos os campos devem ser preenchidos.", parent=popup_window)
                    return
                
                sheet.cell(row=excel_row_index, column=1).value = new_data['nome_epi']
                sheet.cell(row=excel_row_index, column=2).value = new_data['ca']
                sheet.cell(row=excel_row_index, column=3).value = new_data['funcionario']
                sheet.cell(row=excel_row_index, column=4).value = new_data['data_saida']
                sheet.cell(row=excel_row_index, column=5).value = int(new_data['qtd_saida'])

            elif record_type == 'entrada':
                if not all([new_data['mov_type'], new_data['nome_epi'], new_data['data_entrada'], new_data['qtd_entrada']]):
                    messagebox.showwarning("Atenção", "Todos os campos devem ser preenchidos.", parent=popup_window)
                    return

                sheet.cell(row=excel_row_index, column=1).value = new_data['nome_epi']
                sheet.cell(row=excel_row_index, column=2).value = new_data['ca']
                sheet.cell(row=excel_row_index, column=3).value = new_data['mov_type']
                sheet.cell(row=excel_row_index, column=6).value = new_data['data_entrada']
                sheet.cell(row=excel_row_index, column=7).value = int(new_data['qtd_entrada'])

            workbook.save(DB_PATH)
            messagebox.showinfo("Sucesso", "O registro foi atualizado com sucesso.")
            popup_window.destroy()
            self.populate_treeview()

        except ValueError:
            messagebox.showerror("Erro de Formato", "Verifique se a data está em DD/MM/AAAA e a quantidade é um número válido.", parent=popup_window)
        except PermissionError:
            messagebox.showerror("Erro de Permissão", f"Não foi possível salvar as alterações.\n\nPor favor, feche a planilha '{os.path.basename(DB_PATH)}' e tente novamente.")
        except Exception as e:
            messagebox.showerror("Erro Inesperado", f"Ocorreu um erro ao salvar:\n{e}", parent=popup_window)

    def open_epi_edit_popup(self, event):
        selection = self.cadastro_geral_tree.selection()
        if not selection:
            return
        
        selected_item_id = selection[0]
        excel_row_index = int(selected_item_id)
        values = self.cadastro_geral_tree.item(selected_item_id, 'values')
        original_name = values[0]

        popup = ctk.CTkToplevel(self)
        popup.title("Gerenciar EPI")
        popup.transient(self)
        popup.grab_set()
        popup.resizable(False, False)

        main_frame = ctk.CTkFrame(popup, fg_color="transparent")
        main_frame.pack(padx=30, pady=20)

        ctk.CTkLabel(main_frame, text="Gerenciar EPI", font=SUBTITLE_FONT).grid(row=0, column=0, columnspan=2, pady=(0, 20), sticky="ew")

        nome_entry = self.create_form_row(main_frame, "Nome do EPI:", 1)
        ca_entry = self.create_form_row(main_frame, "C.A:", 2)
        preco_entry = self.create_form_row(main_frame, "Preço Unitário (R$):", 3)

        nome_entry.insert(0, values[0])
        ca_entry.insert(0, values[1])
        preco_str = values[2].replace("R$", "").replace(".", "").replace(",", ".").strip()
        preco_entry.insert(0, preco_str)

        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.grid(row=4, column=0, columnspan=2, pady=(20, 0), sticky="ew")

        save_btn = ctk.CTkButton(button_frame, text="Salvar Alterações", font=BUTTON_FONT, fg_color=BUTTON_COLOR, hover_color=BUTTON_HOVER_COLOR,
                                command=lambda: self.save_edited_epi(excel_row_index, nome_entry.get(), ca_entry.get(), preco_entry.get(), popup))
        save_btn.pack(side="left", expand=True, padx=5)

        remove_btn = ctk.CTkButton(button_frame, text="Remover Registro", font=BUTTON_FONT, fg_color=BUTTON_COLOR_RED, hover_color=BUTTON_HOVER_RED,
                                   command=lambda: self.delete_cadastro_epi(excel_row_index, original_name, popup))
        remove_btn.pack(side="left", expand=True, padx=5)

        barcode_btn = ctk.CTkButton(button_frame, text="Gerar Código de Barras", font=BUTTON_FONT, fg_color=BUTTON_COLOR_BLUE, hover_color=BUTTON_HOVER_BLUE,
                                    command=lambda: self.open_barcode_generator_popup(nome_entry.get(), popup))
        barcode_btn.pack(side="left", expand=True, padx=5)

    def save_edited_epi(self, excel_row_index, new_name, new_ca, new_preco_str, popup):
        if not all([new_name, new_ca, new_preco_str]):
            messagebox.showwarning("Atenção", "Todos os campos são obrigatórios.", parent=popup)
            return
        try:
            new_preco = float(new_preco_str.replace(",", "."))
            workbook = self.get_cadastro_epi_workbook()
            sheet = workbook[CADASTRO_WORKSHEET_NAME]

            sheet.cell(row=excel_row_index, column=1).value = new_name
            sheet.cell(row=excel_row_index, column=2).value = new_ca
            sheet.cell(row=excel_row_index, column=3).value = new_preco
            
            workbook.save(DB_PATH)
            messagebox.showinfo("Sucesso", "EPI atualizado com sucesso.", parent=popup)
            popup.destroy()
            self.populate_cadastro_geral_treeview()
            self.populate_controle_epi_options()

        except ValueError:
            messagebox.showerror("Erro de Formato", "O preço deve ser um número válido.", parent=popup)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar: {e}", parent=popup)

    def delete_cadastro_epi(self, excel_row_index, epi_name, popup):
        if not messagebox.askyesno("Confirmar Remoção", f"Tem certeza que deseja remover o EPI '{epi_name}'?\nEsta ação é permanente.", parent=popup):
            return
        try:
            workbook = self.get_cadastro_epi_workbook()
            sheet = workbook[CADASTRO_WORKSHEET_NAME]
            sheet.delete_rows(excel_row_index)
            workbook.save(DB_PATH)
            messagebox.showinfo("Sucesso", f"EPI '{epi_name}' removido.", parent=popup)
            popup.destroy()
            self.populate_cadastro_geral_treeview()
            self.populate_controle_epi_options()
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível remover: {e}", parent=popup)

    def open_barcode_generator_popup(self, epi_name, parent_popup):
        barcode_popup = ctk.CTkToplevel(self)
        barcode_popup.title("Código de Barras")
        barcode_popup.transient(parent_popup)
        barcode_popup.grab_set()
        barcode_popup.resizable(False, False)

        try:
            normalized_name = unicodedata.normalize('NFKD', epi_name).encode('ascii', 'ignore').decode('utf-8').upper()
            if not normalized_name:
                raise ValueError("O nome do EPI resultou em uma string vazia após a normalização.")
            
            code128 = barcode.get_barcode_class('code128')
            barcode_instance = code128(normalized_name, writer=ImageWriter())
            
            buffer = BytesIO()
            barcode_instance.write(buffer)
            buffer.seek(0)
            
            pil_img = Image.open(buffer)
            ctk_img = ctk.CTkImage(light_image=pil_img, size=(pil_img.width, pil_img.height))
            
            img_label = ctk.CTkLabel(barcode_popup, image=ctk_img, text="")
            img_label.pack(padx=20, pady=20)

            button_frame = ctk.CTkFrame(barcode_popup, fg_color="transparent")
            button_frame.pack(padx=20, pady=(0, 20), fill="x")

            save_btn = ctk.CTkButton(button_frame, text="Salvar", command=lambda: self.save_barcode_image(barcode_instance, epi_name, barcode_popup))
            save_btn.pack(side="left", expand=True, padx=10)
            
            cancel_btn = ctk.CTkButton(button_frame, text="Cancelar", command=barcode_popup.destroy, fg_color=BUTTON_COLOR_RED, hover_color=BUTTON_HOVER_RED)
            cancel_btn.pack(side="left", expand=True, padx=10)

        except Exception as e:
            messagebox.showerror("Erro de Código de Barras", f"Não foi possível gerar o código:\n{e}", parent=barcode_popup)
            barcode_popup.destroy()

    def save_barcode_image(self, barcode_instance, epi_name, popup):
        file_path = filedialog.asksaveasfilename(
            initialfile=f"{epi_name}.png",
            defaultextension=".png",
            filetypes=[("PNG files", "*.png"), ("All files", "*.*")]
        )
        if file_path:
            try:
                barcode_instance.write(file_path)
                messagebox.showinfo("Sucesso", f"Código de barras salvo em:\n{file_path}", parent=popup)
                popup.destroy()
            except Exception as e:
                messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar o arquivo:\n{e}", parent=popup)

    def apply_filters(self, _=None):
        self.generate_chart(self.chart_filter.get())

    def clear_all_dashboard_filters(self):
        self.drilldown_employee = None
        self.start_date_cal.delete(0, 'end')
        self.end_date_cal.delete(0, 'end')
        self.dashboard_epi_filter.set("Todos")
        self.apply_filters()

    def update_dashboard_controls(self, _=None):
        chart_type = self.chart_filter.get()
        if chart_type == "Frequência":
            self.view_mode_filter.set("📊 Quantidade")
            self.view_mode_filter.configure(state="disabled")
        else:
            self.view_mode_filter.configure(state="normal")
        self.apply_filters()

    def populate_dashboard_filters(self):
        try:
            df = pd.read_excel(DB_PATH, sheet_name=WORKSHEET_NAME)
            epi_list = ["Todos"] + sorted(df['NOME EPI'].dropna().unique().tolist())
            self.dashboard_epi_filter.configure(values=epi_list)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível carregar filtros de EPI:\n{e}")

    def generate_chart(self, chart_type: str):
        if hasattr(self, 'chart_canvas') and self.chart_canvas:
            self.chart_canvas.get_tk_widget().destroy()
        
        for widget in self.chart_frame.winfo_children():
            widget.destroy()

        try:
            df = pd.read_excel(DB_PATH, sheet_name=WORKSHEET_NAME)
            if 'NOME EPI' in df.columns: df['NOME EPI'] = df['NOME EPI'].astype(str).str.strip()
            if 'FUNCIONARIO' in df.columns: df['FUNCIONARIO'] = df['FUNCIONARIO'].astype(str).str.strip()
            df['DATA SAIDA'] = pd.to_datetime(df['DATA SAIDA'], dayfirst=True, errors='coerce')
            
            df['PREÇO UNITÁRIO'] = pd.to_numeric(df['PREÇO UNITÁRIO'], errors='coerce').fillna(0)
            df['QUANTIDADE SAIDA'] = pd.to_numeric(df['QUANTIDADE SAIDA'], errors='coerce').fillna(0)
            df['VALOR SAIDA'] = df['QUANTIDADE SAIDA'] * df['PREÇO UNITÁRIO']

            selected_epi = self.dashboard_epi_filter.get()
            view_mode = self.view_mode_filter.get()
            
            start_date_str = self.start_date_cal.get()
            end_date_str = self.end_date_cal.get()

            if start_date_str and end_date_str:
                try:
                    start_date = datetime.strptime(start_date_str, "%d/%m/%Y")
                    end_date = datetime.strptime(end_date_str, "%d/%m/%Y")
                    mask = (df['DATA SAIDA'] >= start_date) & (df['DATA SAIDA'] <= end_date)
                    df = df.loc[mask]
                except ValueError:
                    messagebox.showerror("Erro de Data", "Formato de data inválido. Use DD/MM/AAAA.")
                    return
            
            format_as_currency = False
            value_data = None

            if chart_type == "Quantidade Total de EPIs":
                if self.drilldown_employee:
                    df = df[df['FUNCIONARIO'] == self.drilldown_employee]
                    df.dropna(subset=['DATA SAIDA'], inplace=True)
                    if df.empty: raise ValueError(f"Não há dados de saída para o funcionário '{self.drilldown_employee}' no período.")
                    
                    if view_mode == "💸 Valor":
                        monthly_data_with_dates = df.set_index('DATA SAIDA').resample('ME')['VALOR SAIDA'].sum()
                        title = f'Valor Mensal Consumido: {self.drilldown_employee}'
                        ylabel = 'Valor Total (R$)'
                        format_as_currency = True
                        self.sum_label_title.configure(text="Somatório Período:")
                        self.mean_label_title.configure(text="Custo Médio:")
                    else:
                        monthly_data_with_dates = df.set_index('DATA SAIDA').resample('ME')['QUANTIDADE SAIDA'].sum()
                        title = f'Consumo Mensal de: {self.drilldown_employee}'
                        ylabel = 'Quantidade Total de EPIs'
                        self.sum_label_title.configure(text="Somatório Período:")
                        self.mean_label_title.configure(text="Média Período:")
                        if view_mode == "📊+💸 Qtd e Valor":
                            value_data_monthly = df.set_index('DATA SAIDA').resample('ME')['VALOR SAIDA'].sum()
                            value_data = value_data_monthly.reindex(monthly_data_with_dates.index)

                    monthly_data_with_dates = monthly_data_with_dates[monthly_data_with_dates > 0]
                    if monthly_data_with_dates.empty: raise ValueError(f"Não há dados de saída para o funcionário '{self.drilldown_employee}' no período.")
                    
                    data = monthly_data_with_dates.copy()
                    data.index = data.index.strftime('%b/%y')
                else:
                    if selected_epi != "Todos": df = df[df['NOME EPI'] == selected_epi]
                    df.dropna(subset=['FUNCIONARIO', 'QUANTIDADE SAIDA'], inplace=True)
                    df = df[df['FUNCIONARIO'].str.strip() != '']
                    
                    if view_mode == "💸 Valor":
                        data = df.groupby('FUNCIONARIO')['VALOR SAIDA'].sum().sort_values(ascending=False)
                        title, ylabel = 'Valor Consumido por Funcionário', 'Valor Total (R$)'
                        format_as_currency = True
                        self.sum_label_title.configure(text="Somatório Período:")
                        self.mean_label_title.configure(text="Custo Médio:")
                    else:
                        data = df.groupby('FUNCIONARIO')['QUANTIDADE SAIDA'].sum().sort_values(ascending=False)
                        title, ylabel = 'Consumo de EPIs por Funcionário', 'Quantidade Total de EPIs'
                        self.sum_label_title.configure(text="Somatório Período:")
                        self.mean_label_title.configure(text="Média Período:")
                        if view_mode == "📊+💸 Qtd e Valor":
                            value_data = df.groupby('FUNCIONARIO')['VALOR SAIDA'].sum().reindex(data.index)

            elif chart_type == "Frequência":
                if self.drilldown_employee:
                    df = df[df['FUNCIONARIO'] == self.drilldown_employee]
                    df.dropna(subset=['DATA SAIDA'], inplace=True)
                    if df.empty: raise ValueError(f"Não há dados de saída para o funcionário '{self.drilldown_employee}' no período.")
                    
                    monthly_data_with_dates = df.set_index('DATA SAIDA').resample('ME').size()
                    monthly_data_with_dates = monthly_data_with_dates[monthly_data_with_dates > 0]
                    if monthly_data_with_dates.empty: raise ValueError(f"Não há dados de saída para o funcionário '{self.drilldown_employee}' no período.")
                    
                    data = monthly_data_with_dates.copy()
                    data.index = data.index.strftime('%b/%y')
                    title = f'Frequência Mensal de Retirada: {self.drilldown_employee}'
                    ylabel = 'Número de Retiradas'
                else:
                    if selected_epi != "Todos": df = df[df['NOME EPI'] == selected_epi]
                    df.dropna(subset=['FUNCIONARIO'], inplace=True)
                    df = df[df['FUNCIONARIO'].str.strip() != '']
                    if df.empty: raise ValueError("Não há dados de funcionários para os filtros selecionados.")
                    data = df['FUNCIONARIO'].value_counts()
                    title, ylabel = 'Frequência de Retirada por Funcionário', 'Número de Retiradas'
                    self.sum_label_title.configure(text="Somatório Período:")
                    self.mean_label_title.configure(text="Média Período:")

            elif chart_type == "EPIs":
                if selected_epi == "Todos":
                    df.dropna(subset=['NOME EPI'], inplace=True)
                    df = df[df['NOME EPI'].str.strip() != '']
                    if df.empty: raise ValueError("Não há dados de EPIs para os filtros selecionados.")
                    data = df['NOME EPI'].value_counts()
                    title, ylabel = 'Frequência de Retirada por Tipo de EPI', 'Número de Retiradas'
                    self.sum_label_title.configure(text="Somatório Período:")
                    self.mean_label_title.configure(text="Média Período:")
                else:
                    df = df[df['NOME EPI'] == selected_epi]
                    df.dropna(subset=['DATA SAIDA', 'QUANTIDADE SAIDA'], inplace=True)
                    if df.empty: raise ValueError(f"Não há dados de saída para o EPI '{selected_epi}' no período.")
                    
                    if view_mode == "💸 Valor":
                        monthly_data_with_dates = df.set_index('DATA SAIDA').resample('ME')['VALOR SAIDA'].sum()
                        title, ylabel = f'Valor Mensal Consumido: {selected_epi}', 'Valor Total (R$)'
                        format_as_currency = True
                        self.sum_label_title.configure(text="Somatório Período:")
                        self.mean_label_title.configure(text="Custo Médio:")
                    else:
                        monthly_data_with_dates = df.set_index('DATA SAIDA').resample('ME')['QUANTIDADE SAIDA'].sum()
                        title, ylabel = f'Consumo Mensal de: {selected_epi}', 'Quantidade Total Entregue'
                        self.sum_label_title.configure(text="Somatório Período:")
                        self.mean_label_title.configure(text="Média Período:")
                        if view_mode == "📊+💸 Qtd e Valor":
                            value_data_monthly = df.set_index('DATA SAIDA').resample('ME')['VALOR SAIDA'].sum()
                            value_data = value_data_monthly.reindex(monthly_data_with_dates.index)
                    
                    monthly_data_with_dates = monthly_data_with_dates[monthly_data_with_dates > 0]
                    if monthly_data_with_dates.empty: raise ValueError(f"Não há dados de saída para o EPI '{selected_epi}' no período.")
                    
                    data = monthly_data_with_dates.copy()
                    data.index = data.index.strftime('%b/%y')
            
            else: return

            fig, ax = plt.subplots(figsize=(12, 7), dpi=100)
            bars = ax.bar(data.index, data.values, color=BUTTON_COLOR, zorder=3)
            ax.set_title(title, fontsize=16, pad=20); ax.set_ylabel(ylabel, fontsize=12); ax.set_xlabel('')
            ax.tick_params(axis='x', rotation=90, labelsize=10); ax.grid(axis='y', linestyle='--', alpha=0.7, zorder=0)
            fig.subplots_adjust(bottom=0.3)
            
            if format_as_currency:
                ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: locale.currency(x, symbol='R$', grouping=True)))

            annot = ax.annotate("", xy=(0,0), xytext=(0,0), textcoords="offset points", ha='center', va='center', fontsize=10, fontweight="bold", color="white", bbox=dict(boxstyle="round,pad=0.3", fc=BUTTON_HOVER_COLOR, ec="none", lw=0, alpha=0.9))
            annot.set_visible(False)

            def hover(event):
                vis = annot.get_visible()
                if event.inaxes == ax:
                    for i, bar in enumerate(bars):
                        if bar.contains(event)[0]:
                            bar_value = bar.get_height()
                            text = ""
                            if view_mode == "📊+💸 Qtd e Valor" and value_data is not None:
                                valor_r = value_data.iloc[i]
                                text = f"Qtd: {int(bar_value)}\nValor: {locale.currency(valor_r, symbol='R$', grouping=True)}"
                            elif format_as_currency:
                                text = locale.currency(bar_value, symbol='R$', grouping=True)
                            else:
                                text = f"{int(bar_value)}"
                            
                            annot.set_text(text)
                            annot.xy = (bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2)
                            annot.set_visible(True)
                            for b in bars: b.set_color(BUTTON_COLOR)
                            bar.set_color(BUTTON_HOVER_COLOR)
                            fig.canvas.draw_idle()
                            return
                if vis:
                    annot.set_visible(False)
                    for b in bars: b.set_color(BUTTON_COLOR)
                    fig.canvas.draw_idle()

            fig.canvas.mpl_connect("motion_notify_event", hover)
            
            if (chart_type == "Quantidade Total de EPIs" or chart_type == "Frequência") and not self.drilldown_employee:
                def on_employee_bar_double_click(event):
                    if event.inaxes == ax and event.dblclick:
                        for i, bar in enumerate(bars):
                            if bar.contains(event)[0]:
                                employee_name = data.index[i]
                                self.drilldown_employee = employee_name
                                self.apply_filters()
                                return
                fig.canvas.mpl_connect('button_press_event', on_employee_bar_double_click)
            
            elif chart_type == "EPIs" and selected_epi == "Todos":
                def on_bar_double_click(event):
                    if event.inaxes == ax and event.dblclick:
                        for i, bar in enumerate(bars):
                            if bar.contains(event)[0]:
                                selected_epi_name = data.index[i]
                                self.dashboard_epi_filter.set(selected_epi_name)
                                self.apply_filters()
                                return
                fig.canvas.mpl_connect('button_press_event', on_bar_double_click)
            
            elif chart_type == "EPIs" and selected_epi != "Todos":
                def on_monthly_bar_double_click(event):
                    if event.inaxes == ax and event.dblclick:
                        for i, bar in enumerate(bars):
                            if bar.contains(event)[0]:
                                month_end_date = monthly_data_with_dates.index[i]
                                month_start_date = month_end_date.replace(day=1)
                                
                                self.chart_filter.set("Quantidade Total de EPIs")
                                self.start_date_cal.set_date(month_start_date)
                                self.end_date_cal.set_date(month_end_date)
                                
                                self.apply_filters()
                                return
                fig.canvas.mpl_connect('button_press_event', on_monthly_bar_double_click)


            self.chart_canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
            self.chart_canvas.draw(); self.chart_canvas.get_tk_widget().pack(side=ctk.TOP, fill=ctk.BOTH, expand=True)
            
            if not data.empty:
                if view_mode == "📊+💸 Qtd e Valor" and value_data is not None:
                    sum_qty = data.sum()
                    mean_qty = data.mean()
                    sum_val = value_data.sum()
                    mean_val = value_data.mean()
                    self.sum_label_title.configure(text="Somatórios:")
                    self.mean_label_title.configure(text="Médias:")
                    self.sum_label.configure(text=f"Qtd: {int(sum_qty)} | Valor: {locale.currency(sum_val, symbol='R$', grouping=True)}")
                    self.mean_label.configure(text=f"Qtd: {mean_qty:.2f} | Valor: {locale.currency(mean_val, symbol='R$', grouping=True)}")
                else:
                    sum_value = data.sum()
                    mean_value = data.mean()
                    if format_as_currency:
                        self.sum_label.configure(text=locale.currency(sum_value, symbol='R$', grouping=True))
                        self.mean_label.configure(text=locale.currency(mean_value, symbol='R$', grouping=True))
                    else:
                        self.sum_label.configure(text=f"{int(sum_value)}")
                        self.mean_label.configure(text=f"{mean_value:.2f}")
            else:
                self.sum_label.configure(text="-")
                self.mean_label.configure(text="-")

        except FileNotFoundError: messagebox.showerror("Erro", f"Arquivo não encontrado: {DB_PATH}")
        except ValueError as ve: ctk.CTkLabel(self.chart_frame, text=str(ve), font=LABEL_FONT).pack()
        except Exception as e: messagebox.showerror("Erro ao Gerar Gráfico", f"Ocorreu um erro inesperado:\n{e}")

    def select_frame_by_name(self, name):
        self.reset_screens()
        
        all_buttons = [self.controle_button, self.cadastro_button, self.cadastro_geral_button, self.inventario_button, self.dashboard_button, self.updates_button]
        frames_map = {"controle_epi": self.controle_epi_frame, "cadastro_epi": self.cadastro_epi_frame, "cadastro_geral_epi": self.cadastro_geral_epi_frame, "inventario_epi": self.inventario_epi_frame, "dashboard": self.dashboard_frame}
        buttons_map = {"controle_epi": self.controle_button, "cadastro_epi": self.cadastro_button, "cadastro_geral_epi": self.cadastro_geral_button, "inventario_epi": self.inventario_button, "dashboard": self.dashboard_button, "updates": self.updates_button}

        for btn in all_buttons: btn.configure(fg_color=BUTTON_COLOR)
        for fr in frames_map.values(): fr.grid_forget()
        
        if name in buttons_map:
            buttons_map[name].configure(fg_color=BUTTON_HOVER_COLOR)
        if name in frames_map:
            frames_map[name].grid(row=0, column=0, sticky="nsew")
        
        if name == "controle_epi":
            self.populate_controle_epi_options()
            self.update_recent_entries_panel()
            self.populate_saldo_geral_treeview()
        if name == "cadastro_epi":
            self.populate_controle_epi_options()
            self.update_recent_movimentacoes_panel()
        if name == "cadastro_geral_epi":
            self.populate_cadastro_geral_treeview()
        if name == "inventario_epi": 
            self.populate_treeview()
        if name == "dashboard":
            self.drilldown_employee = None
            self.populate_dashboard_filters()
            self.chart_filter.set("Frequência")
            self.dashboard_epi_filter.set("Todos")
            self.view_mode_filter.set("📊 Quantidade")
            self.start_date_cal.set_date(datetime(datetime.now().year, 1, 1))
            self.end_date_cal.set_date(datetime.now())
            self.update_dashboard_controls()

    def controle_button_event(self): self.select_frame_by_name("controle_epi")
    def cadastro_button_event(self): self.select_frame_by_name("cadastro_epi")
    def cadastro_geral_button_event(self): self.select_frame_by_name("cadastro_geral_epi")
    def inventario_button_event(self): self.select_frame_by_name("inventario_epi")
    def dashboard_button_event(self): self.select_frame_by_name("dashboard")

if __name__ == "__main__":
    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("green")
    app = App()
    app.mainloop()